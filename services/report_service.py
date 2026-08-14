from io import BytesIO

from flask import (
    request,
    session,
    render_template,
    send_file
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from database.db import get_connection


# =========================================================
# Helpers
# =========================================================

def _get_institution_id():
    return session.get("institution_id")


def _get_role():
    return session.get("role")


def _get_user_id():
    return session.get("user_id")


# =========================================================
# Get Academic Years
# =========================================================

def _get_academic_years(cur, institution_id):

    cur.execute("""
        SELECT
            id,
            year_name,
            start_date,
            end_date,
            is_current

        FROM academic_years

        WHERE
            institution_id = %s
            AND is_active = TRUE

        ORDER BY
            start_date DESC,
            id DESC
    """, (
        institution_id,
    ))

    return cur.fetchall()


# =========================================================
# Get Allowed Classes
# =========================================================

def _get_allowed_classes(
    cur,
    institution_id,
    role,
    user_id
):

    if role == "institution_admin":

        cur.execute("""
            SELECT
                id,
                class_name

            FROM classes

            WHERE
                institution_id = %s
                AND is_active = TRUE

            ORDER BY
                class_name
        """, (
            institution_id,
        ))

    elif role == "staff":

        cur.execute("""
            SELECT DISTINCT
                c.id,
                c.class_name

            FROM classes c

            JOIN staff_classes sc
                ON sc.class_id = c.id

            WHERE
                c.institution_id = %s

                AND sc.institution_id = %s
                AND sc.staff_id = %s
                AND sc.is_active = TRUE

                AND c.is_active = TRUE

            ORDER BY
                c.class_name
        """, (
            institution_id,
            institution_id,
            user_id
        ))

    else:

        return []


    return cur.fetchall()


# =========================================================
# Verify Class Access
# =========================================================

def _class_is_allowed(
    cur,
    class_id,
    institution_id,
    role,
    user_id
):

    if not class_id:
        return True


    if role == "institution_admin":

        cur.execute("""
            SELECT
                id

            FROM classes

            WHERE
                id = %s
                AND institution_id = %s
                AND is_active = TRUE

            LIMIT 1
        """, (
            class_id,
            institution_id
        ))

    elif role == "staff":

        cur.execute("""
            SELECT
                c.id

            FROM classes c

            JOIN staff_classes sc
                ON sc.class_id = c.id

            WHERE
                c.id = %s
                AND c.institution_id = %s

                AND sc.institution_id = %s
                AND sc.staff_id = %s
                AND sc.is_active = TRUE

                AND c.is_active = TRUE

            LIMIT 1
        """, (
            class_id,
            institution_id,
            institution_id,
            user_id
        ))

    else:

        return False


    return cur.fetchone() is not None


# =========================================================
# Central Report
# =========================================================

# =========================================================
# Central Report
# =========================================================

def central_report():

    institution_id = _get_institution_id()
    role = _get_role()
    user_id = _get_user_id()

    # =====================================================
    # Access
    # =====================================================

    if role not in (
        "institution_admin",
        "staff"
    ):
        return "Unauthorized", 403

    conn = get_connection()
    cur = conn.cursor()

    try:

        # =================================================
        # Filter Values
        # =================================================

        academic_year_id = request.args.get(
            "academic_year_id",
            ""
        ).strip()

        month = request.args.get(
            "month",
            ""
        ).strip()

        class_id = request.args.get(
            "class_id",
            ""
        ).strip()

        # =================================================
        # Academic Years
        # =================================================

        academic_years = _get_academic_years(
            cur,
            institution_id
        )

        # =================================================
        # Allowed Classes
        # =================================================

        classes = _get_allowed_classes(
            cur,
            institution_id,
            role,
            user_id
        )

        # =================================================
        # Academic Year Required
        # =================================================

        if not academic_year_id:

            return render_template(
                "reports/central.html",

                academic_years=academic_years,
                classes=classes,

                selected_academic_year=None,
                selected_month=month,
                selected_class=class_id,

                selected_academic_year_data=None,

                report_data=[],

                total_students=0,
                total_css=0,
                average_css=0,
                average_exam=None,
                average_attendance=None
            )

        # =================================================
        # Validate Academic Year
        # =================================================

        cur.execute("""
            SELECT
                id,
                year_name,
                start_date,
                end_date

            FROM academic_years

            WHERE
                id = %s
                AND institution_id = %s
                AND is_active = TRUE

            LIMIT 1
        """, (
            academic_year_id,
            institution_id
        ))

        academic_year = cur.fetchone()

        if not academic_year:

            return render_template(
                "reports/central.html",

                academic_years=academic_years,
                classes=classes,

                selected_academic_year=academic_year_id,
                selected_month=month,
                selected_class=class_id,

                selected_academic_year_data=None,

                report_data=[],

                total_students=0,
                total_css=0,
                average_css=0,
                average_exam=None,
                average_attendance=None,

                error_message="Invalid academic year."
            )

        # =================================================
        # Verify Class
        # =================================================

        if class_id:

            if not _class_is_allowed(
                cur,
                class_id,
                institution_id,
                role,
                user_id
            ):

                return render_template(
                    "reports/central.html",

                    academic_years=academic_years,
                    classes=classes,

                    selected_academic_year=academic_year_id,
                    selected_month=month,
                    selected_class=None,

                    selected_academic_year_data=academic_year,

                    report_data=[],

                    total_students=0,
                    total_css=0,
                    average_css=0,
                    average_exam=None,
                    average_attendance=None,

                    error_message=(
                        "You do not have access to this class."
                    )
                )

        # =================================================
        # Date Range
        # =================================================

        start_date = academic_year["start_date"]
        end_date = academic_year["end_date"]

        # =================================================
        # Month Filter
        # =================================================

        month_condition = ""

        month_number = None

        if month:

            try:

                month_number = int(month)

            except (
                TypeError,
                ValueError
            ):

                month_number = None

            if month_number not in range(1, 13):

                return render_template(
                    "reports/central.html",

                    academic_years=academic_years,
                    classes=classes,

                    selected_academic_year=academic_year_id,
                    selected_month=month,
                    selected_class=class_id,

                    selected_academic_year_data=academic_year,

                    report_data=[],

                    total_students=0,
                    total_css=0,
                    average_css=0,
                    average_exam=None,
                    average_attendance=None,

                    error_message="Invalid month."
                )

            month_condition = """
                AND EXTRACT(
                    MONTH FROM {date_column}
                ) = %s
            """

        # =================================================
        # Class Condition
        # =================================================

        class_condition = ""

        class_params = []

        if class_id:

            class_condition = """
                AND s.class_id = %s
            """

            class_params.append(
                class_id
            )

        # =================================================
        # Central Performance Query
        # =================================================

        query = f"""

            WITH module_points AS (

                -- =================================================
                -- Reading
                -- =================================================

                SELECT
                    rs.student_id,

                    SUM(
                        COALESCE(
                            rs.points,
                            0
                        )
                    ) AS points

                FROM reading_submissions rs

                WHERE
                    rs.institution_id = %s

                    AND rs.status = 'Approved'

                    AND rs.created_at::date
                        BETWEEN %s AND %s

                    {
                        month_condition.replace(
                            "{date_column}",
                            "rs.created_at"
                        )
                        if month_condition
                        else ""
                    }

                GROUP BY
                    rs.student_id


                UNION ALL


                -- =================================================
                -- Writing
                -- =================================================

                SELECT
                    ws.student_id,

                    SUM(
                        COALESCE(
                            ws.points,
                            0
                        )
                    ) AS points

                FROM writing_submissions ws

                WHERE
                    ws.institution_id = %s

                    AND ws.status = 'Approved'

                    AND ws.created_at::date
                        BETWEEN %s AND %s

                    {
                        month_condition.replace(
                            "{date_column}",
                            "ws.created_at"
                        )
                        if month_condition
                        else ""
                    }

                GROUP BY
                    ws.student_id


                UNION ALL


                -- =================================================
                -- Speaking
                -- =================================================

                SELECT
                    sp.student_id,

                    SUM(
                        COALESCE(
                            sp.points,
                            0
                        )
                    ) AS points

                FROM speaking_submissions sp

                WHERE
                    sp.institution_id = %s

                    AND sp.status = 'Approved'

                    AND sp.presentation_date
                        BETWEEN %s AND %s

                    {
                        month_condition.replace(
                            "{date_column}",
                            "sp.presentation_date"
                        )
                        if month_condition
                        else ""
                    }

                GROUP BY
                    sp.student_id


                UNION ALL


                -- =================================================
                -- Publications
                -- =================================================

                SELECT
                    p.student_id,

                    SUM(
                        COALESCE(
                            p.points,
                            0
                        )
                        +
                        COALESCE(
                            p.bonus_points,
                            0
                        )
                    ) AS points

                FROM publications p

                WHERE
                    p.institution_id = %s

                    AND p.status = 'Approved'

                    AND p.publication_date
                        BETWEEN %s AND %s

                    {
                        month_condition.replace(
                            "{date_column}",
                            "p.publication_date"
                        )
                        if month_condition
                        else ""
                    }

                GROUP BY
                    p.student_id


                UNION ALL


                -- =================================================
                -- Language Skills
                -- =================================================

                SELECT
                    ls.student_id,

                    SUM(
                        COALESCE(
                            ls.points,
                            0
                        )
                        +
                        COALESCE(
                            ls.bonus_points,
                            0
                        )
                    ) AS points

                FROM language_skill_assessments ls

                WHERE
                    ls.institution_id = %s

                    AND ls.status = 'Approved'

                    AND ls.created_at::date
                        BETWEEN %s AND %s

                    {
                        month_condition.replace(
                            "{date_column}",
                            "ls.created_at"
                        )
                        if month_condition
                        else ""
                    }

                GROUP BY
                    ls.student_id


                UNION ALL


                -- =================================================
                -- Achievements
                -- =================================================

                SELECT
                    a.student_id,

                    SUM(
                        COALESCE(
                            a.points,
                            0
                        )
                        +
                        COALESCE(
                            a.bonus_points,
                            0
                        )
                    ) AS points

                FROM achievements a

                WHERE
                    a.institution_id = %s

                    AND a.status = 'Approved'

                    AND a.achievement_date
                        BETWEEN %s AND %s

                    {
                        month_condition.replace(
                            "{date_column}",
                            "a.achievement_date"
                        )
                        if month_condition
                        else ""
                    }

                GROUP BY
                    a.student_id


                UNION ALL


                -- =================================================
                -- Paper Presentations
                -- =================================================

                SELECT
                    pp.student_id,

                    SUM(
                        COALESCE(
                            pp.points,
                            0
                        )
                    ) AS points

                FROM paper_presentations pp

                WHERE
                    pp.institution_id = %s

                    AND pp.status = 'Approved'

                    AND pp.created_at::date
                        BETWEEN %s AND %s

                    {
                        month_condition.replace(
                            "{date_column}",
                            "pp.created_at"
                        )
                        if month_condition
                        else ""
                    }

                GROUP BY
                    pp.student_id
            ),


            -- =================================================
            -- Consolidated CSS
            -- =================================================

            consolidated_points AS (

                SELECT
                    student_id,

                    SUM(points) AS css_points

                FROM module_points

                GROUP BY
                    student_id
            ),


            -- =================================================
            -- Exam Data
            -- =================================================

            exam_data AS (

                SELECT
                    em.student_id,

                    SUM(
                        em.mark
                    ) AS obtained_marks,

                    SUM(
                        e.total_mark
                    ) AS possible_marks

                FROM exam_marks em

                JOIN exams e
                    ON e.id = em.exam_id

                WHERE
                    e.institution_id = %s

                    AND e.is_active = TRUE

                    AND e.exam_date
                        BETWEEN %s AND %s

                    {
                        month_condition.replace(
                            "{date_column}",
                            "e.exam_date"
                        )
                        if month_condition
                        else ""
                    }

                GROUP BY
                    em.student_id
            ),


            -- =================================================
            -- Attendance Data
            -- =================================================

            attendance_data AS (

                SELECT
                    a.student_id,

                    COUNT(*) AS total_periods,

                    COUNT(*) FILTER (
                        WHERE
                            a.status IN (
                                'Present',
                                'Late'
                            )
                    ) AS attended_periods

                FROM attendance a

                WHERE
                    a.institution_id = %s

                    AND a.attendance_date
                        BETWEEN %s AND %s

                    {
                        month_condition.replace(
                            "{date_column}",
                            "a.attendance_date"
                        )
                        if month_condition
                        else ""
                    }

                GROUP BY
                    a.student_id
            )


            -- =================================================
            -- Students
            -- =================================================

            SELECT

                s.id,

                s.admission_no,

                s.full_name,

                c.id AS class_id,

                c.class_name,

                COALESCE(
                    cp.css_points,
                    0
                ) AS css_points,

                COALESCE(
                    ed.obtained_marks,
                    0
                ) AS exam_obtained,

                COALESCE(
                    ed.possible_marks,
                    0
                ) AS exam_possible,

                COALESCE(
                    ad.total_periods,
                    0
                ) AS attendance_total,

                COALESCE(
                    ad.attended_periods,
                    0
                ) AS attendance_attended

            FROM students s

            LEFT JOIN classes c
                ON c.id = s.class_id

            LEFT JOIN consolidated_points cp
                ON cp.student_id = s.id

            LEFT JOIN exam_data ed
                ON ed.student_id = s.id

            LEFT JOIN attendance_data ad
                ON ad.student_id = s.id

            WHERE
                s.institution_id = %s

                AND s.is_active = TRUE

                {class_condition}

            ORDER BY
                c.class_name,
                s.full_name
        """

        # =====================================================
        # Query Parameters
        # =====================================================

        params = []

        # =====================================================
        # Module Date Parameters
        # =====================================================

        module_date_sets = [

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            )
        ]

        for item in module_date_sets:

            params.extend(item)

            if month:

                params.append(
                    month_number
                )

        # =====================================================
        # Exam Parameters
        # =====================================================

        params.extend([
            institution_id,
            start_date,
            end_date
        ])

        if month:

            params.append(
                month_number
            )

        # =====================================================
        # Attendance Parameters
        # =====================================================

        params.extend([
            institution_id,
            start_date,
            end_date
        ])

        if month:

            params.append(
                month_number
            )

        # =====================================================
        # Student Parameters
        # =====================================================

        params.append(
            institution_id
        )

        params.extend(
            class_params
        )

        # =====================================================
        # Execute Query
        # =====================================================

        cur.execute(
            query,
            tuple(params)
        )

        rows = cur.fetchall()

        # =====================================================
        # Display Data
        # =====================================================

        report_data = []

        for row in rows:

            css_points = float(
                row["css_points"] or 0
            )

            exam_obtained = float(
                row["exam_obtained"] or 0
            )

            exam_possible = float(
                row["exam_possible"] or 0
            )

            if exam_possible > 0:

                exam_percentage = round(
                    (
                        exam_obtained
                        / exam_possible
                    ) * 100,
                    2
                )

            else:

                exam_percentage = None

            attendance_total = int(
                row["attendance_total"] or 0
            )

            attendance_attended = int(
                row["attendance_attended"] or 0
            )

            if attendance_total > 0:

                attendance_percentage = round(
                    (
                        attendance_attended
                        / attendance_total
                    ) * 100,
                    2
                )

            else:

                attendance_percentage = None

            report_data.append({

                "id":
                    row["id"],

                "admission_no":
                    row["admission_no"],

                "full_name":
                    row["full_name"],

                "class_id":
                    row["class_id"],

                "class_name":
                    row["class_name"]
                    or "Unassigned",

                "css_points":
                    round(
                        css_points,
                        2
                    ),

                "exam_percentage":
                    exam_percentage,

                "attendance_percentage":
                    attendance_percentage
            })

        # =====================================================
        # Summary
        # =====================================================

        total_students = len(
            report_data
        )

        total_css = round(
            sum(
                item["css_points"]
                for item in report_data
            ),
            2
        )

        average_css = (

            round(
                total_css
                / total_students,
                2
            )

            if total_students > 0

            else 0
        )

        exam_values = [

            item["exam_percentage"]

            for item in report_data

            if item["exam_percentage"]
            is not None
        ]

        attendance_values = [

            item["attendance_percentage"]

            for item in report_data

            if item["attendance_percentage"]
            is not None
        ]

        average_exam = (

            round(
                sum(exam_values)
                / len(exam_values),
                2
            )

            if exam_values

            else None
        )

        average_attendance = (

            round(
                sum(attendance_values)
                / len(attendance_values),
                2
            )

            if attendance_values

            else None
        )

        # =====================================================
        # Render
        # =====================================================

        return render_template(
            "reports/central.html",

            academic_years=academic_years,

            classes=classes,

            selected_academic_year=academic_year_id,

            selected_month=month,

            selected_class=class_id,

            selected_academic_year_data=academic_year,

            report_data=report_data,

            total_students=total_students,

            total_css=total_css,

            average_css=average_css,

            average_exam=average_exam,

            average_attendance=average_attendance
        )

    finally:

        cur.close()
        conn.close()
        
# =========================================================
# Central Report PDF
# =========================================================

def central_report_pdf():

    institution_id = _get_institution_id()
    role = _get_role()
    user_id = _get_user_id()


    # =====================================================
    # Access
    # =====================================================

    if role not in (
        "institution_admin",
        "staff"
    ):

        return "Unauthorized", 403


    # =====================================================
    # ReportLab Imports
    # =====================================================

    from io import BytesIO

    from reportlab.lib import colors

    from reportlab.lib.enums import TA_CENTER

    from reportlab.lib.pagesizes import A4

    from reportlab.lib.styles import (
        getSampleStyleSheet,
        ParagraphStyle
    )

    from reportlab.lib.units import mm

    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle
    )


    conn = get_connection()
    cur = conn.cursor()


    try:

        # =================================================
        # Filters
        # =================================================

        academic_year_id = request.args.get(
            "academic_year_id",
            ""
        ).strip()

        month = request.args.get(
            "month",
            ""
        ).strip()

        class_id = request.args.get(
            "class_id",
            ""
        ).strip()


        # =================================================
        # Academic Year
        # =================================================

        if not academic_year_id:

            return "Academic year is required.", 400


        cur.execute("""
            SELECT
                id,
                year_name,
                start_date,
                end_date

            FROM academic_years

            WHERE
                id = %s
                AND institution_id = %s
                AND is_active = TRUE

            LIMIT 1
        """, (
            academic_year_id,
            institution_id
        ))


        academic_year = cur.fetchone()


        if not academic_year:

            return "Invalid academic year.", 400


        start_date = academic_year["start_date"]

        end_date = academic_year["end_date"]


        # =================================================
        # Validate Month
        # =================================================

        month_number = None


        if month:

            try:

                month_number = int(month)

            except (
                TypeError,
                ValueError
            ):

                return "Invalid month.", 400


            if month_number not in range(1, 13):

                return "Invalid month.", 400


        # =================================================
        # Verify Class
        # =================================================

        if class_id:

            if not _class_is_allowed(
                cur,
                class_id,
                institution_id,
                role,
                user_id
            ):

                return (
                    "You do not have access to this class.",
                    403
                )


        # =================================================
        # Class Condition
        # =================================================

        class_condition = ""

        class_params = []


        if class_id:

            class_condition = """
                AND s.class_id = %s
            """

            class_params.append(
                class_id
            )


        # =================================================
        # Month Conditions
        # =================================================

        reading_month = ""

        writing_month = ""

        speaking_month = ""

        publication_month = ""

        language_month = ""

        achievement_month = ""

        paper_month = ""

        exam_month = ""

        attendance_month = ""


        if month_number:

            reading_month = """
                AND EXTRACT(
                    MONTH FROM rs.created_at
                ) = %s
            """

            writing_month = """
                AND EXTRACT(
                    MONTH FROM ws.created_at
                ) = %s
            """

            speaking_month = """
                AND EXTRACT(
                    MONTH FROM sp.presentation_date
                ) = %s
            """

            publication_month = """
                AND EXTRACT(
                    MONTH FROM p.publication_date
                ) = %s
            """

            language_month = """
                AND EXTRACT(
                    MONTH FROM ls.created_at
                ) = %s
            """

            achievement_month = """
                AND EXTRACT(
                    MONTH FROM a.achievement_date
                ) = %s
            """

            paper_month = """
                AND EXTRACT(
                    MONTH FROM pp.created_at
                ) = %s
            """

            exam_month = """
                AND EXTRACT(
                    MONTH FROM e.exam_date
                ) = %s
            """

            attendance_month = """
                AND EXTRACT(
                    MONTH FROM a.attendance_date
                ) = %s
            """


        # =================================================
        # Central Report Query
        # =================================================

        query = f"""

            WITH module_points AS (

                -- Reading

                SELECT
                    rs.student_id,

                    SUM(
                        COALESCE(
                            rs.points,
                            0
                        )
                    ) AS points

                FROM reading_submissions rs

                WHERE
                    rs.institution_id = %s
                    AND rs.status = 'Approved'

                    AND rs.created_at::date
                        BETWEEN %s AND %s

                    {reading_month}

                GROUP BY
                    rs.student_id


                UNION ALL


                -- Writing

                SELECT
                    ws.student_id,

                    SUM(
                        COALESCE(
                            ws.points,
                            0
                        )
                    ) AS points

                FROM writing_submissions ws

                WHERE
                    ws.institution_id = %s
                    AND ws.status = 'Approved'

                    AND ws.created_at::date
                        BETWEEN %s AND %s

                    {writing_month}

                GROUP BY
                    ws.student_id


                UNION ALL


                -- Speaking

                SELECT
                    sp.student_id,

                    SUM(
                        COALESCE(
                            sp.points,
                            0
                        )
                    ) AS points

                FROM speaking_submissions sp

                WHERE
                    sp.institution_id = %s
                    AND sp.status = 'Approved'

                    AND sp.presentation_date
                        BETWEEN %s AND %s

                    {speaking_month}

                GROUP BY
                    sp.student_id


                UNION ALL


                -- Publications

                SELECT
                    p.student_id,

                    SUM(
                        COALESCE(
                            p.points,
                            0
                        )
                        +
                        COALESCE(
                            p.bonus_points,
                            0
                        )
                    ) AS points

                FROM publications p

                WHERE
                    p.institution_id = %s
                    AND p.status = 'Approved'

                    AND p.publication_date
                        BETWEEN %s AND %s

                    {publication_month}

                GROUP BY
                    p.student_id


                UNION ALL


                -- Language Skills

                SELECT
                    ls.student_id,

                    SUM(
                        COALESCE(
                            ls.points,
                            0
                        )
                        +
                        COALESCE(
                            ls.bonus_points,
                            0
                        )
                    ) AS points

                FROM language_skill_assessments ls

                WHERE
                    ls.institution_id = %s
                    AND ls.status = 'Approved'

                    AND ls.created_at::date
                        BETWEEN %s AND %s

                    {language_month}

                GROUP BY
                    ls.student_id


                UNION ALL


                -- Achievements

                SELECT
                    a.student_id,

                    SUM(
                        COALESCE(
                            a.points,
                            0
                        )
                        +
                        COALESCE(
                            a.bonus_points,
                            0
                        )
                    ) AS points

                FROM achievements a

                WHERE
                    a.institution_id = %s
                    AND a.status = 'Approved'

                    AND a.achievement_date
                        BETWEEN %s AND %s

                    {achievement_month}

                GROUP BY
                    a.student_id


                UNION ALL


                -- Paper Presentations

                SELECT
                    pp.student_id,

                    SUM(
                        COALESCE(
                            pp.points,
                            0
                        )
                    ) AS points

                FROM paper_presentations pp

                WHERE
                    pp.institution_id = %s
                    AND pp.status = 'Approved'

                    AND pp.created_at::date
                        BETWEEN %s AND %s

                    {paper_month}

                GROUP BY
                    pp.student_id
            ),


            consolidated_points AS (

                SELECT
                    student_id,

                    SUM(points)
                    AS css_points

                FROM module_points

                GROUP BY
                    student_id
            ),


            exam_data AS (

                SELECT
                    em.student_id,

                    SUM(
                        em.mark
                    ) AS obtained_marks,

                    SUM(
                        e.total_mark
                    ) AS possible_marks

                FROM exam_marks em

                JOIN exams e
                    ON e.id = em.exam_id

                WHERE
                    e.institution_id = %s

                    AND e.is_active = TRUE

                    AND e.exam_date
                        BETWEEN %s AND %s

                    {exam_month}

                GROUP BY
                    em.student_id
            ),


            attendance_data AS (

                SELECT
                    a.student_id,

                    COUNT(*) AS total_periods,

                    COUNT(*) FILTER (
                        WHERE a.status IN (
                            'Present',
                            'Late'
                        )
                    ) AS attended_periods

                FROM attendance a

                WHERE
                    a.institution_id = %s

                    AND a.attendance_date
                        BETWEEN %s AND %s

                    {attendance_month}

                GROUP BY
                    a.student_id
            )


            SELECT

                s.id,

                s.admission_no,

                s.full_name,

                c.class_name,

                COALESCE(
                    cp.css_points,
                    0
                ) AS css_points,

                COALESCE(
                    ed.obtained_marks,
                    0
                ) AS exam_obtained,

                COALESCE(
                    ed.possible_marks,
                    0
                ) AS exam_possible,

                COALESCE(
                    ad.total_periods,
                    0
                ) AS attendance_total,

                COALESCE(
                    ad.attended_periods,
                    0
                ) AS attendance_attended


            FROM students s

            LEFT JOIN classes c
                ON c.id = s.class_id

            LEFT JOIN consolidated_points cp
                ON cp.student_id = s.id

            LEFT JOIN exam_data ed
                ON ed.student_id = s.id

            LEFT JOIN attendance_data ad
                ON ad.student_id = s.id

            WHERE
                s.institution_id = %s

                AND s.is_active = TRUE

                {class_condition}

            ORDER BY
                c.class_name,
                s.full_name
        """


        # =================================================
        # Query Parameters
        # =================================================

        params = []


        module_sets = [

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            ),

            (
                institution_id,
                start_date,
                end_date
            )
        ]


        month_flags = [

            bool(reading_month),

            bool(writing_month),

            bool(speaking_month),

            bool(publication_month),

            bool(language_month),

            bool(achievement_month),

            bool(paper_month)
        ]


        for item, has_month in zip(
            module_sets,
            month_flags
        ):

            params.extend(item)

            if has_month:

                params.append(
                    month_number
                )


        # Exam

        params.extend([
            institution_id,
            start_date,
            end_date
        ])

        if exam_month:

            params.append(
                month_number
            )


        # Attendance

        params.extend([
            institution_id,
            start_date,
            end_date
        ])

        if attendance_month:

            params.append(
                month_number
            )


        # Students

        params.append(
            institution_id
        )

        params.extend(
            class_params
        )


        cur.execute(
            query,
            tuple(params)
        )

        rows = cur.fetchall()
        
        # =================================================
        # Prepare Report Data
        # =================================================

        report_data = []


        for row in rows:

            css_points = float(
                row["css_points"] or 0
            )


            exam_obtained = float(
                row["exam_obtained"] or 0
            )

            exam_possible = float(
                row["exam_possible"] or 0
            )


            if exam_possible > 0:

                exam_percentage = round(
                    (
                        exam_obtained
                        / exam_possible
                    ) * 100,
                    2
                )

            else:

                exam_percentage = None


            attendance_total = int(
                row["attendance_total"] or 0
            )

            attendance_attended = int(
                row["attendance_attended"] or 0
            )


            if attendance_total > 0:

                attendance_percentage = round(
                    (
                        attendance_attended
                        / attendance_total
                    ) * 100,
                    2
                )

            else:

                attendance_percentage = None


            report_data.append({

                "admission_no":
                    row["admission_no"],

                "full_name":
                    row["full_name"],

                "class_name":
                    row["class_name"]
                    or "Unassigned",

                "css_points":
                    round(
                        css_points,
                        2
                    ),

                "exam_percentage":
                    exam_percentage,

                "attendance_percentage":
                    attendance_percentage
            })


        # =================================================
        # Summary
        # =================================================

        total_students = len(
            report_data
        )


        total_css = round(
            sum(
                item["css_points"]
                for item in report_data
            ),
            2
        )


        average_css = (

            round(
                total_css
                / total_students,
                2
            )

            if total_students > 0

            else 0
        )


        exam_values = [

            item["exam_percentage"]

            for item in report_data

            if item["exam_percentage"]
            is not None
        ]


        attendance_values = [

            item["attendance_percentage"]

            for item in report_data

            if item["attendance_percentage"]
            is not None
        ]


        average_exam = (

            round(
                sum(exam_values)
                / len(exam_values),
                2
            )

            if exam_values

            else None
        )


        average_attendance = (

            round(
                sum(attendance_values)
                / len(attendance_values),
                2
            )

            if attendance_values

            else None
        )


        # =================================================
        # Institution Details
        # =================================================

        cur.execute("""
            SELECT
                name,
                logo

            FROM institutions

            WHERE
                id = %s

            LIMIT 1
        """, (
            institution_id,
        ))


        institution = cur.fetchone()


        institution_name = (
            institution["name"]
            if institution
            else "Institution"
        )


        # =================================================
        # PDF Buffer
        # =================================================

        buffer = BytesIO()


        doc = SimpleDocTemplate(

            buffer,

            pagesize=A4,

            rightMargin=15 * mm,

            leftMargin=15 * mm,

            topMargin=15 * mm,

            bottomMargin=15 * mm
        )


        # =================================================
        # Styles
        # =================================================

        styles = getSampleStyleSheet()


        title_style = ParagraphStyle(

            "ReportTitle",

            parent=styles["Title"],

            fontSize=20,

            leading=24,

            alignment=TA_CENTER,

            spaceAfter=6
        )


        subtitle_style = ParagraphStyle(

            "ReportSubtitle",

            parent=styles["Normal"],

            fontSize=10,

            leading=14,

            alignment=TA_CENTER,

            textColor=colors.grey,

            spaceAfter=15
        )


        heading_style = ParagraphStyle(

            "ReportHeading",

            parent=styles["Heading2"],

            fontSize=13,

            leading=16,

            spaceBefore=10,

            spaceAfter=8
        )


        normal_style = ParagraphStyle(

            "ReportNormal",

            parent=styles["Normal"],

            fontSize=9,

            leading=12
        )


        # =================================================
        # PDF Story
        # =================================================

        story = []


        # =================================================
        # Header
        # =================================================

        story.append(
            Paragraph(
                institution_name,
                title_style
            )
        )


        story.append(
            Paragraph(
                "Central Performance Report",
                subtitle_style
            )
        )


        # =================================================
        # Filter Information
        # =================================================

        filter_data = [

            [
                Paragraph(
                    "<b>Academic Year</b>",
                    normal_style
                ),

                Paragraph(
                    str(
                        academic_year["year_name"]
                    ),
                    normal_style
                )
            ],

            [
                Paragraph(
                    "<b>Period</b>",
                    normal_style
                ),

                Paragraph(
                    f"{start_date} to {end_date}",
                    normal_style
                )
            ],

            [
                Paragraph(
                    "<b>Month</b>",
                    normal_style
                ),

                Paragraph(
                    (
                        str(month_number)
                        if month_number
                        else "All Months"
                    ),
                    normal_style
                )
            ],

            [
                Paragraph(
                    "<b>Class</b>",
                    normal_style
                ),

                Paragraph(
                    (
                        str(
                            report_data[0]["class_name"]
                        )
                        if class_id
                        and report_data
                        else (
                            "Selected Class"
                            if class_id
                            else "All Classes"
                        )
                    ),
                    normal_style
                )
            ]
        ]


        filter_table = Table(
            filter_data,
            colWidths=[
                40 * mm,
                120 * mm
            ]
        )


        filter_table.setStyle(
            TableStyle([

                (
                    "BACKGROUND",
                    (0, 0),
                    (0, -1),
                    colors.HexColor(
                        "#f1f5f9"
                    )
                ),

                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.HexColor(
                        "#cbd5e1"
                    )
                ),

                (
                    "INNERGRID",
                    (0, 0),
                    (-1, -1),
                    0.25,
                    colors.HexColor(
                        "#e2e8f0"
                    )
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                )
            ])
        )


        story.append(
            filter_table
        )


        story.append(
            Spacer(
                1,
                12
            )
        )


        # =================================================
        # Summary Cards
        # =================================================

        summary_data = [

            [
                Paragraph(
                    "<b>Total Students</b>",
                    normal_style
                ),

                Paragraph(
                    "<b>Total CSS</b>",
                    normal_style
                ),

                Paragraph(
                    "<b>Average CSS</b>",
                    normal_style
                ),

                Paragraph(
                    "<b>Average Exam</b>",
                    normal_style
                ),

                Paragraph(
                    "<b>Average Attendance</b>",
                    normal_style
                )
            ],

            [
                str(total_students),

                str(total_css),

                str(average_css),

                (
                    f"{average_exam}%"
                    if average_exam is not None
                    else "N/A"
                ),

                (
                    f"{average_attendance}%"
                    if average_attendance is not None
                    else "N/A"
                )
            ]
        ]


        summary_table = Table(
            summary_data,
            colWidths=[
                32 * mm,
                32 * mm,
                32 * mm,
                32 * mm,
                32 * mm
            ]
        )


        summary_table.setStyle(
            TableStyle([

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#e2e8f0"
                    )
                ),

                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, 1),
                    colors.white
                ),

                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.HexColor(
                        "#cbd5e1"
                    )
                ),

                (
                    "INNERGRID",
                    (0, 0),
                    (-1, -1),
                    0.25,
                    colors.HexColor(
                        "#cbd5e1"
                    )
                ),

                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER"
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    7
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    7
                )
            ])
        )


        story.append(
            summary_table
        )


        story.append(
            Spacer(
                1,
                15
            )
        )


        # =================================================
        # Student Performance Table
        # =================================================

        story.append(
            Paragraph(
                "Student Performance",
                heading_style
            )
        )


        table_data = [

            [
                "#",
                "Admission No.",
                "Student",
                "Class",
                "CSS",
                "Exam %",
                "Attendance %"
            ]
        ]


        for index, item in enumerate(
            report_data,
            start=1
        ):

            table_data.append([

                str(index),

                str(
                    item["admission_no"]
                    or "-"
                ),

                Paragraph(
                    str(
                        item["full_name"]
                        or "-"
                    ),
                    normal_style
                ),

                Paragraph(
                    str(
                        item["class_name"]
                    ),
                    normal_style
                ),

                str(
                    item["css_points"]
                ),

                (
                    f'{item["exam_percentage"]}%'
                    if item["exam_percentage"]
                    is not None
                    else "N/A"
                ),

                (
                    f'{item["attendance_percentage"]}%'
                    if item[
                        "attendance_percentage"
                    ] is not None
                    else "N/A"
                )
            ])


        if len(table_data) == 1:

            table_data.append([

                "-",

                "-",

                "No student performance data found.",

                "-",

                "-",

                "-",

                "-"
            ])


        student_table = Table(

            table_data,

            colWidths=[

                8 * mm,

                25 * mm,

                48 * mm,

                30 * mm,

                20 * mm,

                25 * mm,

                30 * mm
            ],

            repeatRows=1
        )


        student_table.setStyle(
            TableStyle([

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#1e293b"
                    )
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),

                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    8
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor(
                        "#cbd5e1"
                    )
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),

                (
                    "ALIGN",
                    (0, 0),
                    (1, -1),
                    "CENTER"
                ),

                (
                    "ALIGN",
                    (4, 1),
                    (-1, -1),
                    "CENTER"
                ),

                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor(
                            "#f8fafc"
                        )
                    ]
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                )
            ])
        )


        story.append(
            student_table
        )


        # =================================================
        # Build PDF
        # =================================================

        doc.build(
            story
        )


        buffer.seek(0)


        # =================================================
        # Return PDF
        # =================================================

        from flask import send_file


        filename = (
            "central_performance_report.pdf"
        )


        return send_file(

            buffer,

            as_attachment=True,

            download_name=filename,

            mimetype="application/pdf"
        )


    finally:

        cur.close()

        conn.close()
        

# =========================================================
# Central Report - Excel Export
# =========================================================

def central_report_excel():

    institution_id = session.get("institution_id")
    role = session.get("role")
    user_id = session.get("user_id")

    # =====================================================
    # Access
    # =====================================================

    if role not in (
        "institution_admin",
        "staff"
    ):
        return "Unauthorized", 403

    conn = get_connection()
    cur = conn.cursor()

    try:

        # =================================================
        # Filters
        # =================================================

        academic_year_id = request.args.get(
            "academic_year_id",
            ""
        ).strip()

        month = request.args.get(
            "month",
            ""
        ).strip()

        class_id = request.args.get(
            "class_id",
            ""
        ).strip()

        # =================================================
        # Academic Year Required
        # =================================================

        if not academic_year_id:
            return "Academic year is required.", 400

        # =================================================
        # Validate Academic Year
        # =================================================

        cur.execute("""
            SELECT
                id,
                year_name,
                start_date,
                end_date

            FROM academic_years

            WHERE
                id = %s
                AND institution_id = %s
                AND is_active = TRUE

            LIMIT 1
        """, (
            academic_year_id,
            institution_id
        ))

        academic_year = cur.fetchone()

        if not academic_year:
            return "Invalid academic year.", 400

        # =================================================
        # Verify Class Access
        # =================================================

        if class_id:

            if not _class_is_allowed(
                cur,
                class_id,
                institution_id,
                role,
                user_id
            ):
                return "You do not have access to this class.", 403

        # =================================================
        # Date Range
        # =================================================

        start_date = academic_year["start_date"]
        end_date = academic_year["end_date"]

        # =================================================
        # Month Validation
        # =================================================

        month_number = None

        if month:

            try:
                month_number = int(month)

            except (TypeError, ValueError):
                return "Invalid month.", 400

            if month_number not in range(1, 13):
                return "Invalid month.", 400

        # =================================================
        # Month Conditions
        # =================================================

        def month_condition(date_column):

            if month_number is not None:

                return f"""
                    AND EXTRACT(
                        MONTH FROM {date_column}
                    ) = %s
                """

            return ""

        # =================================================
        # Class Condition
        # =================================================

        class_condition = ""

        if class_id:

            class_condition = """
                AND s.class_id = %s
            """

        # =================================================
        # Central Query
        # =================================================

        query = f"""
            WITH module_points AS (

                SELECT
                    rs.student_id,
                    SUM(
                        COALESCE(
                            rs.points,
                            0
                        )
                    ) AS points

                FROM reading_submissions rs

                WHERE
                    rs.institution_id = %s
                    AND rs.status = 'Approved'
                    AND rs.created_at::date
                        BETWEEN %s AND %s
                    {month_condition("rs.created_at")}

                GROUP BY
                    rs.student_id


                UNION ALL


                SELECT
                    ws.student_id,
                    SUM(
                        COALESCE(
                            ws.points,
                            0
                        )
                    ) AS points

                FROM writing_submissions ws

                WHERE
                    ws.institution_id = %s
                    AND ws.status = 'Approved'
                    AND ws.created_at::date
                        BETWEEN %s AND %s
                    {month_condition("ws.created_at")}

                GROUP BY
                    ws.student_id


                UNION ALL


                SELECT
                    sp.student_id,
                    SUM(
                        COALESCE(
                            sp.points,
                            0
                        )
                    ) AS points

                FROM speaking_submissions sp

                WHERE
                    sp.institution_id = %s
                    AND sp.status = 'Approved'
                    AND sp.presentation_date
                        BETWEEN %s AND %s
                    {month_condition("sp.presentation_date")}

                GROUP BY
                    sp.student_id


                UNION ALL


                SELECT
                    p.student_id,
                    SUM(
                        COALESCE(
                            p.points,
                            0
                        )
                        +
                        COALESCE(
                            p.bonus_points,
                            0
                        )
                    ) AS points

                FROM publications p

                WHERE
                    p.institution_id = %s
                    AND p.status = 'Approved'
                    AND p.publication_date
                        BETWEEN %s AND %s
                    {month_condition("p.publication_date")}

                GROUP BY
                    p.student_id


                UNION ALL


                SELECT
                    ls.student_id,
                    SUM(
                        COALESCE(
                            ls.points,
                            0
                        )
                        +
                        COALESCE(
                            ls.bonus_points,
                            0
                        )
                    ) AS points

                FROM language_skill_assessments ls

                WHERE
                    ls.institution_id = %s
                    AND ls.status = 'Approved'
                    AND ls.created_at::date
                        BETWEEN %s AND %s
                    {month_condition("ls.created_at")}

                GROUP BY
                    ls.student_id


                UNION ALL


                SELECT
                    a.student_id,
                    SUM(
                        COALESCE(
                            a.points,
                            0
                        )
                        +
                        COALESCE(
                            a.bonus_points,
                            0
                        )
                    ) AS points

                FROM achievements a

                WHERE
                    a.institution_id = %s
                    AND a.status = 'Approved'
                    AND a.achievement_date
                        BETWEEN %s AND %s
                    {month_condition("a.achievement_date")}

                GROUP BY
                    a.student_id


                UNION ALL


                SELECT
                    pp.student_id,
                    SUM(
                        COALESCE(
                            pp.points,
                            0
                        )
                    ) AS points

                FROM paper_presentations pp

                WHERE
                    pp.institution_id = %s
                    AND pp.status = 'Approved'
                    AND pp.created_at::date
                        BETWEEN %s AND %s
                    {month_condition("pp.created_at")}

                GROUP BY
                    pp.student_id
            ),


            consolidated_points AS (

                SELECT
                    student_id,
                    SUM(points) AS css_points

                FROM module_points

                GROUP BY
                    student_id
            ),


            exam_data AS (

                SELECT
                    em.student_id,

                    SUM(
                        em.mark
                    ) AS obtained_marks,

                    SUM(
                        e.total_mark
                    ) AS possible_marks

                FROM exam_marks em

                JOIN exams e
                    ON e.id = em.exam_id

                WHERE
                    e.institution_id = %s
                    AND e.is_active = TRUE
                    AND e.exam_date
                        BETWEEN %s AND %s
                    {month_condition("e.exam_date")}

                GROUP BY
                    em.student_id
            ),


            attendance_data AS (

                SELECT
                    a.student_id,

                    COUNT(*) AS total_periods,

                    COUNT(*) FILTER (
                        WHERE
                            a.status IN (
                                'Present',
                                'Late'
                            )
                    ) AS attended_periods

                FROM attendance a

                WHERE
                    a.institution_id = %s
                    AND a.attendance_date
                        BETWEEN %s AND %s
                    {month_condition("a.attendance_date")}

                GROUP BY
                    a.student_id
            )


            SELECT

                s.id,

                s.admission_no,

                s.full_name,

                c.id AS class_id,

                c.class_name,

                COALESCE(
                    cp.css_points,
                    0
                ) AS css_points,

                COALESCE(
                    ed.obtained_marks,
                    0
                ) AS exam_obtained,

                COALESCE(
                    ed.possible_marks,
                    0
                ) AS exam_possible,

                COALESCE(
                    ad.total_periods,
                    0
                ) AS attendance_total,

                COALESCE(
                    ad.attended_periods,
                    0
                ) AS attendance_attended

            FROM students s

            LEFT JOIN classes c
                ON c.id = s.class_id

            LEFT JOIN consolidated_points cp
                ON cp.student_id = s.id

            LEFT JOIN exam_data ed
                ON ed.student_id = s.id

            LEFT JOIN attendance_data ad
                ON ad.student_id = s.id

            WHERE
                s.institution_id = %s
                AND s.is_active = TRUE

                {class_condition}

            ORDER BY
                c.class_name,
                s.full_name
        """

        # =================================================
        # Query Parameters
        # =================================================

        params = []

        module_sets = [
            (
                institution_id,
                start_date,
                end_date
            ),
            (
                institution_id,
                start_date,
                end_date
            ),
            (
                institution_id,
                start_date,
                end_date
            ),
            (
                institution_id,
                start_date,
                end_date
            ),
            (
                institution_id,
                start_date,
                end_date
            ),
            (
                institution_id,
                start_date,
                end_date
            ),
            (
                institution_id,
                start_date,
                end_date
            )
        ]

        for item in module_sets:

            params.extend(item)

            if month_number is not None:
                params.append(month_number)

        # =================================================
        # Exam Parameters
        # =================================================

        params.extend([
            institution_id,
            start_date,
            end_date
        ])

        if month_number is not None:
            params.append(month_number)

        # =================================================
        # Attendance Parameters
        # =================================================

        params.extend([
            institution_id,
            start_date,
            end_date
        ])

        if month_number is not None:
            params.append(month_number)

        # =================================================
        # Student Parameters
        # =================================================

        params.append(
            institution_id
        )

        if class_id:
            params.append(
                class_id
            )

        # =================================================
        # Execute Query
        # =================================================

        cur.execute(
            query,
            tuple(params)
        )

        rows = cur.fetchall()
        
        # =================================================
        # Prepare Excel Data
        # =================================================

        report_data = []

        for row in rows:

            css_points = float(
                row["css_points"] or 0
            )

            exam_obtained = float(
                row["exam_obtained"] or 0
            )

            exam_possible = float(
                row["exam_possible"] or 0
            )

            if exam_possible > 0:

                exam_percentage = round(
                    (
                        exam_obtained
                        / exam_possible
                    ) * 100,
                    2
                )

            else:

                exam_percentage = None

            attendance_total = int(
                row["attendance_total"] or 0
            )

            attendance_attended = int(
                row["attendance_attended"] or 0
            )

            if attendance_total > 0:

                attendance_percentage = round(
                    (
                        attendance_attended
                        / attendance_total
                    ) * 100,
                    2
                )

            else:

                attendance_percentage = None

            report_data.append({

                "admission_no":
                    row["admission_no"],

                "full_name":
                    row["full_name"],

                "class_name":
                    row["class_name"]
                    or "Unassigned",

                "css_points":
                    round(
                        css_points,
                        2
                    ),

                "exam_percentage":
                    exam_percentage,

                "attendance_percentage":
                    attendance_percentage
            })

        # =================================================
        # Summary
        # =================================================

        total_students = len(
            report_data
        )

        total_css = round(
            sum(
                item["css_points"]
                for item in report_data
            ),
            2
        )

        average_css = (

            round(
                total_css
                / total_students,
                2
            )

            if total_students > 0

            else 0
        )

        exam_values = [

            item["exam_percentage"]

            for item in report_data

            if item["exam_percentage"]
            is not None
        ]

        attendance_values = [

            item["attendance_percentage"]

            for item in report_data

            if item["attendance_percentage"]
            is not None
        ]

        average_exam = (

            round(
                sum(exam_values)
                / len(exam_values),
                2
            )

            if exam_values

            else None
        )

        average_attendance = (

            round(
                sum(attendance_values)
                / len(attendance_values),
                2
            )

            if attendance_values

            else None
        )

        # =================================================
        # Create Workbook
        # =================================================

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Central Report"

        # =================================================
        # Styles
        # =================================================

        title_font = Font(
            bold=True,
            size=16
        )

        header_font = Font(
            bold=True
        )

        bold_font = Font(
            bold=True
        )

        center_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        thin_side = Side(
            style="thin"
        )

        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )

        # =================================================
        # Title
        # =================================================

        worksheet.merge_cells(
            "A1:G1"
        )

        worksheet["A1"] = (
            "Central Performance Report"
        )

        worksheet["A1"].font = title_font

        worksheet["A1"].alignment = center_alignment

        # =================================================
        # Report Information
        # =================================================

        worksheet["A3"] = "Academic Year"

        worksheet["B3"] = (
            academic_year["year_name"]
        )

        worksheet["D3"] = "Month"

        if month_number is not None:

            month_names = [
                "",
                "January",
                "February",
                "March",
                "April",
                "May",
                "June",
                "July",
                "August",
                "September",
                "October",
                "November",
                "December"
            ]

            worksheet["E3"] = (
                month_names[month_number]
            )

        else:

            worksheet["E3"] = "All Months"

        worksheet["A4"] = "Class"

        if class_id:

            cur.execute("""
                SELECT
                    class_name

                FROM classes

                WHERE
                    id = %s
                    AND institution_id = %s

                LIMIT 1
            """, (
                class_id,
                institution_id
            ))

            selected_class_row = (
                cur.fetchone()
            )

            if selected_class_row:

                worksheet["B4"] = (
                    selected_class_row["class_name"]
                )

            else:

                worksheet["B4"] = "Selected Class"

        else:

            worksheet["B4"] = "All Classes"

        # =================================================
        # Summary
        # =================================================

        worksheet["A6"] = "Total Students"

        worksheet["B6"] = total_students

        worksheet["C6"] = "Total CSS Points"

        worksheet["D6"] = total_css

        worksheet["E6"] = "Average CSS"

        worksheet["F6"] = average_css

        worksheet["A7"] = "Average Exam"

        worksheet["B7"] = (
            average_exam
            if average_exam is not None
            else "No Data"
        )

        worksheet["C7"] = "Average Attendance"

        worksheet["D7"] = (
            average_attendance
            if average_attendance is not None
            else "No Data"
        )

        for cell in (
            "A6",
            "C6",
            "E6",
            "A7",
            "C7"
        ):

            worksheet[cell].font = bold_font

        # =================================================
        # Table Header
        # =================================================

        header_row = 9

        headers = [
            "#",
            "Admission No.",
            "Student",
            "Class",
            "CSS Points",
            "Exam %",
            "Attendance %"
        ]

        for column_number, header in enumerate(
            headers,
            start=1
        ):

            cell = worksheet.cell(
                row=header_row,
                column=column_number
            )

            cell.value = header

            cell.font = header_font

            cell.alignment = center_alignment

            cell.border = border

        # =================================================
        # Student Rows
        # =================================================

        current_row = header_row + 1

        for index, student in enumerate(
            report_data,
            start=1
        ):

            values = [
                index,
                student["admission_no"],
                student["full_name"],
                student["class_name"],
                student["css_points"],
                (
                    student["exam_percentage"]
                    if student["exam_percentage"]
                    is not None
                    else "No Data"
                ),
                (
                    student["attendance_percentage"]
                    if student["attendance_percentage"]
                    is not None
                    else "No Data"
                )
            ]

            for column_number, value in enumerate(
                values,
                start=1
            ):

                cell = worksheet.cell(
                    row=current_row,
                    column=column_number
                )

                cell.value = value

                cell.border = border

                if column_number in (
                    1,
                    5,
                    6,
                    7
                ):

                    cell.alignment = center_alignment

            current_row += 1

        # =================================================
        # Column Widths
        # =================================================

        column_widths = {
            "A": 8,
            "B": 18,
            "C": 30,
            "D": 20,
            "E": 15,
            "F": 15,
            "G": 18
        }

        for column, width in column_widths.items():

            worksheet.column_dimensions[
                column
            ].width = width

        # =================================================
        # Freeze Header
        # =================================================

        worksheet.freeze_panes = "A10"

        # =================================================
        # Save Workbook to Memory
        # =================================================

        output = BytesIO()

        workbook.save(
            output
        )

        output.seek(0)

        # =================================================
        # Filename
        # =================================================

        filename = (
            "central_performance_report.xlsx"
        )

        # =================================================
        # Send Excel File
        # =================================================

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

    finally:

        cur.close()
        conn.close()                


